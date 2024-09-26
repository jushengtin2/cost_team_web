from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
import os
from io import BytesIO
import shutil
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side
import re
app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "http://localhost:3000"}})  #讓3000來的請求都通過CORS 之後架server會需要改
CORS(app, resources={r"/*": {"origins": "http://taiwan-cost-team"}})  #測試點
CORS(app, resources={r"/*": {"origins": "http://taiwan-cost-team.auth.hpicorp.net"}})  #測試點
CORS(app, resources={r"/*": {"origins": "http://15.38.111.74:3000"}})  #測試點
CORS(app, resources={r"/*": {"origins": "http://15.38.111.74"}})  #測試點
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 50 MB
UPLOAD_FOLDER = './uploads'
 
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

file_names = {}  #存上傳的檔案 因為上傳的program matrix不會真的只叫做"program matrix" 所以用一個字典來存檔名

@app.route('/', methods=['GET'])
def get_data():
    return jsonify({"message": "Hello from Flask!"})

@app.route('/NB_upload', methods=['POST'])
def upload_file():
    global file_names

    # 獲取上傳的文件
    program_matrix_file = request.files.get('programMatrixFile')
    mspeke_file = request.files.get('mspekeFile')
    hardware_qual_matrix_file = request.files.get('hardwareQualMatrixFile')
    CPC_file = request.files.get('CPCFile')

    
    if program_matrix_file:
        program_matrix_path = os.path.join(UPLOAD_FOLDER, program_matrix_file.filename)
        program_matrix_file.save(program_matrix_path)  # 如果檢查通過，保存文件
        file_names['program_matrix_file'] = program_matrix_file.filename
    if mspeke_file:
        mspeke_file_path = os.path.join(UPLOAD_FOLDER, mspeke_file.filename)
        mspeke_file.save(mspeke_file_path)
        file_names['mspeke_file'] = mspeke_file.filename
    if hardware_qual_matrix_file:
        hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, hardware_qual_matrix_file.filename)
        hardware_qual_matrix_file.save(hardware_qual_matrix_path)
        file_names['hardware_qual_matrix_file'] = hardware_qual_matrix_file.filename
    if CPC_file:
        cpc_path = os.path.join(UPLOAD_FOLDER, CPC_file.filename)
        CPC_file.save(cpc_path)
        file_names['CPC_file'] = CPC_file.filename

    try:
        # 檢查 Program Matrix 文件
        if program_matrix_file:
            with pd.ExcelFile(program_matrix_file) as xls:
                if 'Program Matrix' not in xls.sheet_names:
                    raise ValueError("Sheet 'Program Matrix' not found")
                print('2')    
                df = pd.read_excel(xls, sheet_name='Program Matrix', skiprows=4)
                program_matrix_headers = list(df.columns)  # 檢查標題

                expected_program_matrix_headers = [
                    'Category / Manufacturing Comments',
                    'Release(s)',
                    'Description',
                    'AV\nLevel 2', 
                    'SA\nLevel 3',     
                    'Component\nLevel 4',  
                ]

                # 測試 2: 檢查標題是否匹配
                actual_headers_filtered = [header for header in program_matrix_headers if 'Unnamed' not in header]

                # Check if the key headers match the expected ones
                if expected_program_matrix_headers != actual_headers_filtered[:len(expected_program_matrix_headers)]:
                    raise ValueError("Headers in Program Matrix do not match the expected values.")
                print('3')  

        # 檢查 Mspeke 文件
        if mspeke_file:
            with pd.ExcelFile(mspeke_file) as xls:
                if 'HW' not in xls.sheet_names:
                    raise ValueError("Sheet 'HW' not found")
                df = pd.read_excel(xls, sheet_name='HW', skiprows=4)
                mspeke_headers = list(df.columns)  # 檢查標題
                if 'Feature Full Name' not in mspeke_headers or 'Notes' not in mspeke_headers:
                    raise ValueError("'Feature Full Name' or 'Notes' not found in Mspeke headers")

            # 檢查 Hardware Qual Matrix 文件
        if hardware_qual_matrix_file:
            with pd.ExcelFile(hardware_qual_matrix_file) as xls:
                df = pd.read_excel(xls, skiprows=1)
                hqm_headers = list(df.columns)  # 檢查標題
                if 'HP Part No.' not in hqm_headers:
                    raise ValueError("'HP Part No.' not found in HQM headers")

        # 返回成功訊息
        return jsonify({"message": "Files uploaded and validated successfully"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/NB_delete', methods=['POST'])
def delete_folder():
    try:
        if os.path.exists(UPLOAD_FOLDER):
            shutil.rmtree(UPLOAD_FOLDER)  # 刪除整個資料夾及其內容
            os.makedirs(UPLOAD_FOLDER)    # 刪除後重新創建空的資料夾，方便之後的上傳操作
            
            return jsonify({"message": "資料夾已成功刪除"}), 200
        else:
            return jsonify({"error": "資料夾不存在"}), 400
    except Exception as e:
        return jsonify({"error": f"刪除資料夾過程中發生錯誤: {str(e)}"}), 500

@app.route('/NB_bom_cost_check', methods=['GET'])  
def bom_cost_check():
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))

    if not os.path.exists(program_matrix_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400

    try:
        with pd.ExcelFile(program_matrix_path) as xls:      #使用 with 語句確保文件被關閉!!!! 因為如果不是用with檔案就會一直讓檔案是被操作狀態 會導致之後要按delete時出現 E-busy error

            df = pd.read_excel(xls, sheet_name='Program Matrix', skiprows=4)
            
            program_matrix_headers = list(df.columns)  # Extract the headers

            # Define key headers
            av_header = 'AV\nLevel 2'
            sa_header = 'SA\nLevel 3'
            component_header = 'Component\nLevel 4'
            
            #重新命名COLUMN 因為錢的地方TITLE會是空的 所以要加上COLUMN(且column會多個av sa可能性!)
            av_count = 1
            sa_count = 1
            new_columns = {}
            
            for i in range(len(program_matrix_headers)):
                col_name = program_matrix_headers[i]
                
                if av_header in col_name:
                    av_count = 1  # Reset counter when AV header is found
                    
                if sa_header in col_name:
                    sa_count = 1  # Reset counter when SA header is found
                    
                # Rename columns between 'AV\nLevel 2' and 'SA\nLevel 3' 如果很多site會多個av可能性
                if av_header in program_matrix_headers and program_matrix_headers.index(av_header) < i < program_matrix_headers.index(sa_header):
                    if 'Unnamed' in col_name:
                        new_columns[col_name] = f'av_price{av_count}'
                        av_count += 1

                # Rename columns between 'SA\nLevel 3' and 'Component\nLevel 4' 如果很多site會多個sa可能性
                if sa_header in program_matrix_headers and program_matrix_headers.index(sa_header) < i < program_matrix_headers.index(component_header):
                    if 'Unnamed' in col_name:
                        new_columns[col_name] = f'sa_price{sa_count}'
                        sa_count += 1
            
            # Apply the renaming
            df.rename(columns=new_columns, inplace=True)

            df['Group'] = df['Release(s)'].ffill()

            # Find the indices where 'Release(s)' is not empty (non-null and not empty strings)
            block_start_indices = df.index[df['Release(s)'].notna() & (df['Release(s)'] != '')].tolist()

            error_result = []

            for start, end in zip(block_start_indices, block_start_indices[1:] + [None]):
                block = df.iloc[start:end]
                error_result_temp = []  # Initialize a flat list for results

                # Add block metadata
                error_result_temp.append(block.iloc[0]['AV\nLevel 2'])
                error_result_temp.append(block.iloc[0]['Description'])
                
                # Function to compute total differences
                def calculate_difference(av_column, sa_column):
                    if av_column in block.columns and sa_column in block.columns:
                        total_av = block[av_column].sum()
                        total_sa = block[sa_column].sum()
                        total_diff = total_av - total_sa
                        if abs(total_diff) > 0.01:
                            return total_diff
                    return " "

                # Append calculated differences directly to the error_result_temp list
                error_result_temp.append(calculate_difference('av_price1', 'sa_price1'))
                error_result_temp.append(calculate_difference('av_price2', 'sa_price2'))
                error_result_temp.append(calculate_difference('av_price3', 'sa_price3'))
                error_result_temp.append(calculate_difference('av_price4', 'sa_price4'))

                # Append the flat list to error_result
                if any(item != " " for item in error_result_temp[2:]):  # Check only the difference elements
                    error_result.append(error_result_temp)
                    
        wb = Workbook() #創建工作簿並將結果寫入到指定的工作表中
        ws = wb.active
        ws.title = 'BOM Cost Check'

        ws.append(['AV Level 2', 'Description', 'AV1 Price Difference', 'AV2 Price Difference', 'AV3 Price Difference', 'AV4 Price Difference'])    #寫入header

        for row in error_result:    #寫入數據
            ws.append(row)

        
        excel_buffer = BytesIO()    #保存工作簿到BytesIO對象
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()                  #確保工作簿已關閉

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='BOM_cost_check.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/NB_bom_cost_check_for_highlight_file', methods=['GET'])
def bom_cost_check_for_highlight_file():
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))

    if not os.path.exists(program_matrix_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400

    try:
        # Load the original workbook to apply color formatting
        wb = load_workbook(program_matrix_path)
        ws = wb['Program Matrix']

        with pd.ExcelFile(program_matrix_path) as xls:
            sheet_names = xls.sheet_names
            df = pd.read_excel(xls, sheet_name='Program Matrix', skiprows=4)

            program_matrix_headers = list(df.columns)

            # Define key headers
            av_header = 'AV\nLevel 2'
            sa_header = 'SA\nLevel 3'
            component_header = 'Component\nLevel 4'

            # Renaming logic
            av_count = 1
            sa_count = 1
            new_columns = {}
            
            for i in range(len(program_matrix_headers)):
                col_name = program_matrix_headers[i]
                
                if av_header in col_name:
                    av_count = 1
                    
                if sa_header in col_name:
                    sa_count = 1
                    
                if av_header in program_matrix_headers and program_matrix_headers.index(av_header) < i < program_matrix_headers.index(sa_header):
                    if 'Unnamed' in col_name:
                        new_columns[col_name] = f'av_price{av_count}'
                        av_count += 1

                if sa_header in program_matrix_headers and program_matrix_headers.index(sa_header) < i < program_matrix_headers.index(component_header):
                    if 'Unnamed' in col_name:
                        new_columns[col_name] = f'sa_price{sa_count}'
                        sa_count += 1
            
            df.rename(columns=new_columns, inplace=True)

            df['Group'] = df['Release(s)'].ffill()
            block_start_indices = df.index[df['Release(s)'].notna() & (df['Release(s)'] != '')].tolist()
            error_result = []

            # To store the positions of the cells that need to be colored
            cells_to_color = []

            for start, end in zip(block_start_indices, block_start_indices[1:] + [None]):
                block = df.iloc[start:end]
                error_result_temp = []

                error_result_temp.append(block.iloc[0]['AV\nLevel 2'])
                error_result_temp.append(block.iloc[0]['Description'])
                
                def calculate_difference(av_column, sa_column):
                    if av_column in block.columns and sa_column in block.columns:
                        total_av = block[av_column].sum()
                        total_sa = block[sa_column].sum()
                        total_diff = total_av - total_sa
                        if abs(total_diff) > 0.01:
                            # Find the cell coordinates to color
                            av_cell = block[av_column].iloc[0]
                            av_row_index = block.index[0]
                            cells_to_color.append((av_column, av_row_index + 1)) # add 1 to row index to match Excel's 1-based indexing
                            return total_diff
                        else:
                            return " "  # No significant difference
                    return " "  # Columns not found
  

                error_result_temp.append(calculate_difference('av_price1', 'sa_price1'))
                error_result_temp.append(calculate_difference('av_price2', 'sa_price2'))
                error_result_temp.append(calculate_difference('av_price3', 'sa_price3'))
                error_result_temp.append(calculate_difference('av_price4', 'sa_price4'))

                if any(item != " " for item in error_result_temp[2:]):
                    error_result.append(error_result_temp)
                    
        print(cells_to_color)

        # Apply color formatting to the original Excel file
        fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")
        for col, idx in cells_to_color:
            # Convert column name to Excel column letter and adjust for header rows (skiprows)
            col_index = df.columns.get_loc(col) + 1
            col_letter = openpyxl.utils.get_column_letter(col_index)
            # The row index in ws needs to account for the 5 header rows skipped when reading the data (skiprows=4)
            ws[f"{col_letter}{idx + 5}"].fill = fill

        # Save the modified workbook to a buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='BOM_cost_check_color.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/NB_hqm_based_component_check', methods=['GET'])
def hqm_based_component_check():
    global file_names
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))
    mspeke_path = os.path.join(UPLOAD_FOLDER, file_names.get('mspeke_file', ''))
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('hardware_qual_matrix_file', ''))
    print(file_names)

    if not (os.path.exists(program_matrix_path)):
        print(f"Error: {program_matrix_path} 不存在")
    if not (os.path.exists(mspeke_path)):
        print(f"Error: {mspeke_path} 不存在")
    if not (os.path.exists(hardware_qual_matrix_path)):
        print(f"Error: {hardware_qual_matrix_path} 不存在")
    
    try:
        with pd.ExcelFile(program_matrix_path) as pm, pd.ExcelFile(mspeke_path) as mspeke, pd.ExcelFile(hardware_qual_matrix_path) as hqm:
            pm_sheet_names = pm.sheet_names
            mspeke_sheet_names = mspeke.sheet_names

            df_pm = pd.read_excel(pm, sheet_name=pm_sheet_names[1], skiprows=4)
            df_mspeke = pd.read_excel(mspeke, sheet_name=mspeke_sheet_names[1], skiprows=4)
            df_hqm = pd.read_excel(hqm, skiprows=1, usecols=['ID','HP Part No.', 'Qual Status'])

            start_idx = df_pm[df_pm['Category / Manufacturing Comments'] == 'DIB Hardware'].index[0]
            end_idx = df_pm[(df_pm.index > start_idx) & (df_pm['Category / Manufacturing Comments'].notna())].index[0]

            df_pm = df_pm.drop(df_pm.index[start_idx:end_idx]).reset_index(drop=True)
            df_mspeke = df_mspeke.iloc[:, [1, 4, 8]].dropna(subset=[df_mspeke.columns[4]])
            
            sections = {}
            current_label = None

            for index, row in df_pm.iterrows():
                if pd.notna(row['SA\nLevel 3']):
                    current_label = row['SA\nLevel 3']
                    sections[current_label] = []
                if current_label:
                    sections[current_label].append(row)

            for label in sections:
                sections[label] = pd.DataFrame(sections[label])

            df_hqm['BOM Description'] = None
            df_hqm['MSPEKE Description'] = None

            for index, row in df_hqm.iterrows():
                if row['HP Part No.'] and row['Qual Status']:
                    hqm_part_number = row['HP Part No.']
                    hqm_qual_status = row['Qual Status']
                    for key, value in sections.items():
                        if hqm_part_number in value['Component\nLevel 4'].values:
                            component_value = value[value['Component\nLevel 4'] == hqm_part_number]
                            component_description = component_value['Description'].values[0]
                            df_hqm.at[index, 'BOM Description'] = component_description
                            
                            max_ratio = 0
                            max_mspeke_item = None
                            for idx, mspeke_item in df_mspeke.iterrows():
                                ratio = smith_waterman(component_description, mspeke_item['Feature Full Name'])
                                if max_ratio < ratio:
                                    max_ratio = ratio
                                    max_mspeke_item = mspeke_item['Feature Full Name']
                            df_hqm.at[index, 'MSPEKE Description'] = max_mspeke_item
                            break

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_hqm.to_excel(writer, index=False, sheet_name='Component Error List')

            # 獲取當前活頁簿和工作表
            workbook = writer.book
            worksheet = workbook.active

            # 定義綠色和淺咖啡色填充樣式
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            light_brown_fill = PatternFill(start_color="D3B58C", end_color="D3B58C", fill_type="solid")

            # 遍歷 df_hqm 的每一行，根據條件設置顏色
            for row_idx, row in enumerate(df_hqm.itertuples(), start=2):  # 開始於第2行（Excel中的行從1開始）
                if pd.isna(row._2):  # '_2' 對應的是 'HP Part No.' 列
                    if pd.notna(row.ID) and not any(char.isdigit() for char in str(row.ID)):
                        # ID 欄位沒有數字時，設置綠色
                        for col_idx in range(1, 4): #只要HQM raw data塗色就好
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = green_fill
                    else:
                        # ID 欄位包含數字時，設置淺咖啡色
                        for col_idx in range(1, 4):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = light_brown_fill

        output.seek(0)

        # 返回文件作為下載
        return send_file(
            output,
            as_attachment=True,
            download_name='HQM_Based_component_error_list.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/NB_bom_based_component_check', methods=['GET'])
def bom_based_component_check():
    global file_names
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))
    mspeke_path = os.path.join(UPLOAD_FOLDER, file_names.get('mspeke_file', ''))
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('hardware_qual_matrix_file', ''))


    if not (os.path.exists(program_matrix_path) and os.path.exists(mspeke_path) and os.path.exists(hardware_qual_matrix_path)):
        return jsonify({"error": "One or more files are missing. Please upload the files first."}), 400
    
    try:
        # 讀取 Excel 文件
        with pd.ExcelFile(program_matrix_path) as pm, pd.ExcelFile(mspeke_path) as mspeke, pd.ExcelFile(hardware_qual_matrix_path) as hqm:
            pm_sheet_names = pm.sheet_names
            mspeke_sheet_names = mspeke.sheet_names

            df_pm = pd.read_excel(pm, sheet_name=pm_sheet_names[1], skiprows=4)
            df_mspeke = pd.read_excel(mspeke, sheet_name=mspeke_sheet_names[1], skiprows=4)
            df_hqm = pd.read_excel(hqm, skiprows=1)

            # 清理和準備數據(去除DIB部分)
            start_idx = df_pm[df_pm['Category / Manufacturing Comments'] == 'DIB Hardware'].index[0]  
            end_idx = df_pm[(df_pm.index > start_idx) & (df_pm['Category / Manufacturing Comments'].notna())].index[0]
            df_pm = df_pm.drop(df_pm.index[start_idx:end_idx]).reset_index(drop=True)

            #選出Mspeke需要的部分
            df_mspeke = df_mspeke.iloc[:, [0, 1, 4, 8]].dropna(subset=[df_mspeke.columns[1]])
            df_hqm = df_hqm[['HP Part No.', 'Qual Status']].dropna(subset=['HP Part No.'])
            hqm_dict = df_hqm.set_index('HP Part No.')['Qual Status'].to_dict()

            # 創建新列來存儲結果
            df_pm['Max_MSPEKE_Item'] = None
            df_pm['Max_MSPEKE_Item_Note'] = None
            df_pm['HQM status'] = None
            
            for index, row in df_pm.iterrows():
                if pd.notna(row['Component\nLevel 4']):  #For RCTO也需要對照mspeke
                    cmp_level4 = row['Description']
                    if 'lbl' not in cmp_level4.lower() and 'doc' not in cmp_level4.lower() and 'icon' not in cmp_level4.lower(): #把不需要對照的刪掉
                        max_ratio = 0
                        max_mspeke_item = None
                        max_mspeke_item_note = None

                        cmp_level4_array = [item.strip() for item in re.split(r'[,\s]+', cmp_level4) if item]
                        if 'IC' in cmp_level4_array:
                            for _, item in df_mspeke.iterrows():
                                if item['Feature Category'] == 'IC / Sensory / Controller' or item['Feature Category'] == 'Processor':
                                    cmpji3_level4_numbers = re.findall(r'\d+', cmp_level4)
                                    cmp_level4_numbers_str = ''.join(cmpji3_level4_numbers)
                                    ratio = smith_waterman(cmp_level4_numbers_str, item['Feature Full Name']) #我發現IC如果只取數字來比可能會比較好
                                    if ratio > max_ratio:
                                        max_ratio = ratio
                                        max_mspeke_item = item['Feature Full Name']
                                        max_mspeke_item_note = item['Notes']
                        elif 'CONN' in cmp_level4_array or 'CON' in cmp_level4_array or 'CN' in cmp_level4_array:
                            for _, item in df_mspeke.iterrows():
                                if item['Feature Category'] == 'Connectors' or item['Feature Category'] == 'Power Cord':
                                    ratio = smith_waterman(cmp_level4, item['Feature Full Name'])
                                    if ratio > max_ratio:
                                        max_ratio = ratio  
                                        max_mspeke_item = item['Feature Full Name']
                                        max_mspeke_item_note = item['Notes']
                        else:
                            for _, item in df_mspeke.iterrows():
                                ratio = smith_waterman(cmp_level4, item['Feature Full Name'])
                                if ratio > max_ratio:
                                    max_ratio = ratio
                                    max_mspeke_item = item['Feature Full Name']
                                    max_mspeke_item_note = item['Notes']
                    

                        # 插入結果到對應的行
                        df_pm.at[index, 'Max_MSPEKE_Item'] = max_mspeke_item
                        df_pm.at[index, 'Max_MSPEKE_Item_Note'] = max_mspeke_item_note
                        df_pm.at[index, 'Description'] = ' '*4 +df_pm.at[index, 'Description'] #所 
                        #print(f'1. {cmp_level4}, 2.{max_mspeke_item} , 3. {max_ratio}')

                        if row['Component\nLevel 4'] not in hqm_dict:
                            df_pm.at[index, 'HQM status'] = 'Not in Hardware Qual Matrix'
                        else:
                            qual_status_value = hqm_dict[row['Component\nLevel 4']]
                            df_pm.at[index, 'HQM status'] = qual_status_value
                if pd.notna(row['SA\nLevel 3']):
                    df_pm.at[index, 'Description'] = ' '*2 +df_pm.at[index, 'Description']
                if pd.notna(row['Component\nLevel 5']):
                    df_pm.at[index, 'Description'] = ' '*6 +df_pm.at[index, 'Description']
                
                
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_pm.to_excel(writer, index=False, sheet_name='BOM Based Component Check')

            # 獲取工作表
            worksheet = writer.sheets['BOM Based Component Check']

            # 定義顏色
            pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 藍色填充
            header_yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 標題黃色填充

            # 定義邊框樣式
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # 使用列名來填充顏色
            max_mspeke_item_col = df_pm.columns.get_loc('Max_MSPEKE_Item') + 1  # +1 是因為 openpyxl 列索引從 1 開始
            max_mspeke_item_note_col = df_pm.columns.get_loc('Max_MSPEKE_Item_Note') + 1
            hqm_status_col = df_pm.columns.get_loc('HQM status') + 1

            # 先將不包含特定列的標題行（第一行）塗上黃色
            for col_idx in range(1, worksheet.max_column + 1):
                if col_idx not in [max_mspeke_item_col, max_mspeke_item_note_col, hqm_status_col]:
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_yellow_fill
                    cell.border = thin_border

            # 塗上藍色
            for row_idx in range(2, worksheet.max_row + 1):  # 從第2行開始（跳過標題行）
                category_comments = worksheet.cell(row=row_idx, column=df_pm.columns.get_loc('Category / Manufacturing Comments') + 1).value
                release_s = worksheet.cell(row=row_idx, column=df_pm.columns.get_loc('Release(s)') + 1).value
                description = worksheet.cell(row=row_idx, column=df_pm.columns.get_loc('Description') + 1).value

                if category_comments and (release_s is None or release_s == '') and (description is None or description == ''):
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = blue_fill
                        cell.border = thin_border
            # 塗上粉紅色和黃色並保持邊框
            for cell in worksheet[worksheet.cell(row=1, column=max_mspeke_item_col).column_letter]:
                cell.fill = pink_fill
                cell.border = thin_border

            for cell in worksheet[worksheet.cell(row=1, column=max_mspeke_item_note_col).column_letter]:
                cell.fill = pink_fill
                cell.border = thin_border

            for cell in worksheet[worksheet.cell(row=1, column=hqm_status_col).column_letter]:
                cell.fill = yellow_fill
                cell.border = thin_border 
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='BOM_Based_component_error_list.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/CPC_check', methods=['GET'])
def CPC_check():

    CPC_path = os.path.join(UPLOAD_FOLDER, file_names.get('CPC_file', ''))
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))

    if not os.path.exists(program_matrix_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400
    if not os.path.exists(CPC_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400
      
    try:
        with pd.ExcelFile(CPC_path) as CPC, pd.ExcelFile(program_matrix_path) as BOM: 
            optional_sheets = [sheet for sheet in CPC.sheet_names if sheet.startswith('OptionSA')]
            if len(optional_sheets) < 2:
                return jsonify({"error": "The required OptionSA sheets are missing."}), 400

            df_OptionSA = pd.read_excel(CPC, sheet_name=optional_sheets[0])
            df_OptionSA_SUM = pd.read_excel(CPC, sheet_name=optional_sheets[1])

            SA_PN_dict = {}
            OptionSA = []
            OptionSA_SUM = []
            OptionSA_error_result = []
            OptionSA_SUM_error_result = []

            # 紀錄OptionSA的價錢總和
            for index, row in df_OptionSA.iterrows():
                key = f"{row['SA PN']}_{row['Program Matrix']}"
                OptionSA.append(key)
                if key not in SA_PN_dict:
                    SA_PN_dict[key] = row['Cost']
                else:
                    SA_PN_dict[key] += row['Cost']
            
            OptionSA = tuple(set(OptionSA))

            # 計算OptionSA_SUM跟OptionSA加起來是否一樣 還有是否 只有出現在OptionSA_SUM但沒有在OptionSA (會儲存在第二個sheet)
            for index, row in df_OptionSA_SUM.iterrows():
                key = f"{row['SA PN']}_{row['Program Matrix']}"
                OptionSA_SUM.append(key)
                if key not in SA_PN_dict and key!='':
                    OptionSA_SUM_error_result.append([row['SA PN'], row['Program Matrix'], '', 'OptionSA_SUM Key not found in OptionSA'])
                elif abs(row['Cost'] - SA_PN_dict[key]) > 0.01:  # Allow for small floating point differences
                    OptionSA_SUM_error_result.append([row['SA PN'], row['Program Matrix'], 'Error Cost', ''])

            # 找出出現在OptionSA但沒有在OptionSA_SUM(會儲存在第一個sheet)
            for key in OptionSA:
                if key not in OptionSA_SUM:
                    OptionSA_error_result.append([key])
        
            ##########開始對照BOM的錢跟CPC SA是否一樣#########

            df_BOM = pd.read_excel(BOM, sheet_name='Program Matrix', skiprows=4)
            program_matrix_headers = list(df_BOM.columns)  # Extract the headers
            # Define key headers
            av_header = 'AV\nLevel 2'
            sa_header = 'SA\nLevel 3'
            component_header = 'Component\nLevel 4'
            
            #重新命名COLUMN 因為錢的地方TITLE會是空的 所以要加上COLUMN(且column會多個av sa可能性!)
            av_count = 1
            sa_count = 1
            new_columns = {}
            
            for i in range(len(program_matrix_headers)):
                col_name = program_matrix_headers[i]
                
                if av_header in col_name:
                    av_count = 1  # Reset counter when AV header is found
                    
                if sa_header in col_name:
                    sa_count = 1  # Reset counter when SA header is found
                    
                # Rename columns between 'AV\nLevel 2' and 'SA\nLevel 3' 如果很多site會多個av可能性
                if av_header in program_matrix_headers and program_matrix_headers.index(av_header) < i < program_matrix_headers.index(sa_header):
                    if 'Unnamed' in col_name:
                        new_columns[col_name] = f'av_price{av_count}'
                        av_count += 1

                # Rename columns between 'SA\nLevel 3' and 'Component\nLevel 4' 如果很多site會多個sa可能性
                if sa_header in program_matrix_headers and program_matrix_headers.index(sa_header) < i < program_matrix_headers.index(component_header):
                    if 'Unnamed' in col_name:
                        new_columns[col_name] = f'sa_price{sa_count}'
                        sa_count += 1 
            # Apply the renaming
            df_BOM.rename(columns=new_columns, inplace=True)

            df_BOM = df_BOM[['SA\nLevel 3','sa_price1']].dropna(subset=['sa_price1']).drop_duplicates() #先假設BPM只有一個SA(可能到時候會很多SA)
            sa_dict = dict(zip(df_BOM['SA\nLevel 3'], df_BOM['sa_price1']))
            
            CPC_BOM_COST_ERROR = []

            for index, row in df_OptionSA_SUM.iterrows():  #找在OptionSA_SUM的零件是否錢跟BOM同

                if row['SA PN'] not in sa_dict:
                    CPC_BOM_COST_ERROR.append([f'{row["SA PN"]} not in BOM', row['Program Matrix'], '', ''])
                elif abs(row['Cost'] - sa_dict[row['SA PN']]) >0.01:
                    CPC_BOM_COST_ERROR.append([row['SA PN'], row['Program Matrix'], row['Cost'], sa_dict[row['SA PN']]])

        # Create Excel workbook
        wb = Workbook()

        # First sheet: OptionSA Check
        ws = wb.active
        ws.title = 'OptionSA Check'
        ws.append(['Exist in OptionSA_SUM or not'])
        for row in OptionSA_error_result:
            ws.append(row)

        # Second sheet: OptionSA_SUM Check
        ws_second = wb.create_sheet(title='OptionSA_SUM Check')
        ws_second.append(['OptionSA_SUM SA PN', 'Program Matrix', 'OptionSA - OptionSA_SUM Cost Difference', 'Exist or Not'])
        for row in OptionSA_SUM_error_result:
            ws_second.append(row)

        # Second sheet: OptionSA_SUM Check
        ws_third = wb.create_sheet(title='BOM and OptionSA_SUM Check')
        ws_third.append(['OptionSA_SUM SA PN','Program Matrix' ,'OptionSA_SUM $$', 'BOM $$'])
        for row in CPC_BOM_COST_ERROR:
            ws_third.append(row)

        # Save workbook to BytesIO object
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()  # Ensure the workbook is closed

        # Return the Excel file as a downloadable attachment
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='CPC_check.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    


#Smith-Waterman algorithm計算文字相似度
def smith_waterman( seq1, seq2, match_score=2, mismatch_score=-1, gap_score=-1):  
    len1, len2 = len(seq1), len(seq2)
    score_matrix = np.zeros((len1 + 1, len2 + 1), dtype=int)
    traceback_matrix = np.zeros((len1 + 1, len2 + 1), dtype=int)

    max_score = 0
    max_pos = None
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            match = score_matrix[i-1][j-1] + (match_score if seq1[i-1] == seq2[j-1] else mismatch_score)
            delete = score_matrix[i-1][j] + gap_score
            insert = score_matrix[i][j-1] + gap_score
            score_matrix[i][j] = max(0, match, delete, insert)
            if score_matrix[i][j] == 0:
                traceback_matrix[i][j] = 0
            elif score_matrix[i][j] == match:
                traceback_matrix[i][j] = 1
            elif score_matrix[i][j] == delete:
                traceback_matrix[i][j] = 2
            elif score_matrix[i][j] == insert:
                traceback_matrix[i][j] = 3

            if score_matrix[i][j] >= max_score:
                max_score = score_matrix[i][j]
                max_pos = (i, j)
    
    aligned_seq1, aligned_seq2 = [], []
    i, j = max_pos
    while traceback_matrix[i][j] != 0:
        if traceback_matrix[i][j] == 1:
            aligned_seq1.append(seq1[i-1])
            aligned_seq2.append(seq2[j-1])
            i -= 1
            j -= 1
        elif traceback_matrix[i][j] == 2:
            aligned_seq1.append(seq1[i-1])
            aligned_seq2.append('-')
            i -= 1
        elif traceback_matrix[i][j] == 3:
            aligned_seq1.append('-')
            aligned_seq2.append(seq2[j-1])
            j -= 1

    return max_score


if __name__ == '__main__':
    app.run(debug=True,host='0.0.0.0', port=8080)
