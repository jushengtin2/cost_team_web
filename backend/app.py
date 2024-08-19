from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
import os
from io import BytesIO

import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "http://localhost:3000"}})  #讓3000來的請求都通過CORS 之後架server會需要改

UPLOAD_FOLDER = './uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

file_names = {}  #存上傳的檔案 因為上傳的program matrix不會真的只叫做"program matrix" 所以用一個字典來存檔名

@app.route('/', methods=['GET'])
def get_data():
    return jsonify({"message": "Hello from Flask!"})

@app.route('/upload', methods=['POST'])
def upload_file():
    global file_names

    program_matrix_file = request.files.get('programMatrixFile')
    mspeke_file = request.files.get('mspekeFile')
    hardware_qual_matrix_file = request.files.get('hardwareQualMatrixFile')

    if not program_matrix_file or not mspeke_file or not hardware_qual_matrix_file:
        return jsonify({"error": "Missing file"}), 400

    program_matrix_path = os.path.join(UPLOAD_FOLDER, program_matrix_file.filename)
    mspeke_file_path = os.path.join(UPLOAD_FOLDER, mspeke_file.filename)
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, hardware_qual_matrix_file.filename)

    program_matrix_file.save(program_matrix_path)
    mspeke_file.save(mspeke_file_path)
    hardware_qual_matrix_file.save(hardware_qual_matrix_path)

    # 记录文件名
    file_names['program_matrix_file'] = program_matrix_file.filename
    file_names['mspeke_file'] = mspeke_file.filename
    file_names['hardware_qual_matrix_file'] = hardware_qual_matrix_file.filename

    print("已上傳了:", file_names)  # 確定有存到上傳的檔案

    return jsonify({"message": "已成功上傳跟保存"}), 200


@app.route('/delete', methods=['POST'])
def delete_folder():
    try:
        if os.path.exists(UPLOAD_FOLDER):
            shutil.rmtree(UPLOAD_FOLDER)  # 刪除整個資料夾及其內容
            os.makedirs(UPLOAD_FOLDER)    # 刪除後重新創建空的資料夾，方便之後的上傳操作
            
            return jsonify({"message": "資料夾已成功刪除"}), 200
        else:
            return jsonify({"error": "資料夾不存在"}), 400
    except Exception as e:
        return jsonify({"error": f"刪除資料夾過程中發生錯誤: {str(e)}"}), 500

@app.route('/bom_cost_check', methods=['GET'])
def bom_cost_check():
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))

    if not os.path.exists(program_matrix_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400

    try:
        with pd.ExcelFile(program_matrix_path) as xls:      #使用 with 語句確保文件被關閉!!!! 因為如果不是用with檔案就會一直讓檔案是被操作狀態 會導致之後要按delete時出現 E-busy error
            sheet_names = xls.sheet_names
            df = pd.read_excel(xls, sheet_name=sheet_names[1], skiprows=4)  # sheet_names[1] is Program Matrix
        
        df.rename(columns={'Unnamed: 4': 'av_price', 'Unnamed: 6': 'sa_price'}, inplace=True)
        df['Group'] = df['Release(s)'].ffill()

        block_start_indices = df.index[df['Release(s)'] == 'All (NPI 2024)'].tolist()

        error_result = []
        error_result_AV = []

        for start, end in zip(block_start_indices, block_start_indices[1:] + [None]):
            block = df.iloc[start:end]
            
            if 'av_price' in block.columns and 'sa_price' in block.columns:
                total_av_price = block['av_price'].sum()
                total_sa_price = block['sa_price'].sum()
                total_different_price = total_av_price - total_sa_price

                if abs(total_different_price) > 0.01:
                    error_result.append([block.iloc[0]['AV\nLevel 2'], block.iloc[0]['Description'], total_av_price, total_sa_price, total_different_price])
                    error_result_AV.append(block.iloc[0]['AV\nLevel 2'])

        # 創建工作簿並將結果寫入到指定的工作表中
        wb = Workbook()
        ws = wb.active
        ws.title = 'BOM Cost Check'

        # 寫入header
        ws.append(['AV Level 2', 'Description', 'Total AV Price', 'Total SA Price', 'Price Difference'])

        # 寫入數據
        for row in error_result:
            ws.append(row)

        # 保存工作簿到BytesIO對象
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()  # 確保工作簿已關閉

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='program_matrix_highlight_color.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500





@app.route('/hqm_based_component_check', methods=['GET'])
def hqm_based_component_check():
    global file_names
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))
    mspeke_path = os.path.join(UPLOAD_FOLDER, file_names.get('mspeke_file', ''))
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('hardware_qual_matrix_file', ''))
    print(file_names)

    if not (os.path.exists(program_matrix_path)):
        print(f"Error: {program_matrix_path} 不存在")
    if not (os.path.exists(mspeke_path)):
        print(f"Error: {mspeke_path} 不存在")
    if not (os.path.exists(hardware_qual_matrix_path)):
        print(f"Error: {hardware_qual_matrix_path} 不存在")
    
    try:
        with pd.ExcelFile(program_matrix_path) as pm, pd.ExcelFile(mspeke_path) as mspeke:
            pm_sheet_names = pm.sheet_names
            mspeke_sheet_names = mspeke.sheet_names

            df_pm = pd.read_excel(pm, sheet_name=pm_sheet_names[1], skiprows=4)
            df_mspeke = pd.read_excel(mspeke, sheet_name=mspeke_sheet_names[1], skiprows=4)
            df_hqm = pd.read_excel(hardware_qual_matrix_path, skiprows=1)
        
        df_mspeke = df_mspeke.iloc[:, [1, 4, 8]].dropna()
        df_hqm = df_hqm['HP Part No.'].dropna()

        sections = {}
        current_label = None
        result = {}

        for index, row in df_pm.iterrows():
            if pd.notna(row['SA\nLevel 3']):
                current_label = row['SA\nLevel 3']
                sections[current_label] = []
            if current_label:
                sections[current_label].append(row)

        for label in sections:
            sections[label] = pd.DataFrame(sections[label])

        for hqm_part_number in df_hqm:
            result[hqm_part_number] = []
            HAS_FOUNDED = False
            for key, value in sections.items():
                if hqm_part_number in value['Component\nLevel 4'].values:
                    HAS_FOUNDED = True
                    component_value = value[value['Component\nLevel 4'] == hqm_part_number]
                    component_description = component_value['Description'].values[0]
                    component_qty = value.iloc[0]['Quantity']

                    result[hqm_part_number].extend([component_qty, component_description])

            if not HAS_FOUNDED:
                result[hqm_part_number].append('Cannot find this component in Program Matrix')

        final_result = []

        for key, value in result.items():
            max_ratio = 0
            max_mspeke_item = None
            if len(value) > 1:
                for idx, mspeke_item in df_mspeke.iterrows():
                    ratio = smith_waterman(value[1], mspeke_item['Feature Full Name'])
                    if max_ratio < ratio:
                        max_ratio = ratio
                        max_mspeke_item = mspeke_item
                final_result.append([
                    key, value[1], value[0], max_mspeke_item['Feature Full Name'], max_mspeke_item['Notes']
                ])
            else:
                final_result.append([key, value[0]])

        wb = Workbook()
        ws = wb.active
        ws.title = 'HQM Based Component Check'

        ws.append([
            'Component Part Number', 'Program Matrix Description', 'Program Matrix Qty',
            'PM Description -> MSPEKE Description', 'PM Description -> MSPEKE Notes'
        ])

        for row in final_result:
            ws.append(row)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='HQM_Based_component_error_list.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/bom_based_component_check', methods=['GET'])
def bom_based_component_check():
    global file_names
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))
    mspeke_path = os.path.join(UPLOAD_FOLDER, file_names.get('mspeke_file', ''))
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('hardware_qual_matrix_file', ''))

    if not (os.path.exists(program_matrix_path) and os.path.exists(mspeke_path) and os.path.exists(hardware_qual_matrix_path)):
        return jsonify({"error": "One or more files are missing. Please upload the files first."}), 400

    try:
        with pd.ExcelFile(program_matrix_path) as pm, pd.ExcelFile(mspeke_path) as mspeke:
            pm_sheet_names = pm.sheet_names
            mspeke_sheet_names = mspeke.sheet_names

            df_pm = pd.read_excel(pm, sheet_name=pm_sheet_names[1], skiprows=4)
            df_mspeke = pd.read_excel(mspeke, sheet_name=mspeke_sheet_names[1], skiprows=4)
            df_hqm = pd.read_excel(hardware_qual_matrix_path, skiprows=1)

        start_idx = df_pm[df_pm['Category / Manufacturing Comments'] == 'DIB Hardware'].index[0]
        end_idx = df_pm[(df_pm.index > start_idx) & (df_pm['Category / Manufacturing Comments'].notna())].index[0]
        df_pm = df_pm.drop(df_pm.index[start_idx:end_idx]).reset_index(drop=True)

        df_mspeke = df_mspeke[['Feature\nID', 'Feature Full Name', 'Notes']].dropna(subset=['Notes'])
        df_hqm = df_hqm['HP Part No.'].dropna()

        sections = {}
        not_in_hqm_list = []
        result = {}

        current_label = None
        for index, row in df_pm.iterrows():
            if pd.notna(row['SA\nLevel 3']):
                current_label = row['SA\nLevel 3']
                sections[current_label] = []
            if current_label:
                sections[current_label].append(row)

        for label in sections:
            sections[label] = pd.DataFrame(sections[label])

        for key, value in sections.items():
            for _, row in value.iterrows():
                if pd.notna(row['Component\nLevel 4']):
                    if row['Component\nLevel 4'] not in df_hqm.values:
                        not_in_hqm_list.append(row['Component\nLevel 4'])

                    max_ratio = 0
                    max_mspeke_item = None
                    max_mspeke_item_note = None
                    for _, item in df_mspeke.iterrows():
                        if pd.notna(item['Feature Full Name']):
                            ratio = smith_waterman(row['Description'], item['Feature Full Name'])
                            if ratio > max_ratio:
                                max_ratio = ratio
                                max_mspeke_item = item['Feature Full Name']
                                max_mspeke_item_note = item['Notes']

                    result[row['Component\nLevel 4']] = [
                        row['Description'],
                        max_mspeke_item,
                        value.iloc[0]['Quantity'],
                        max_mspeke_item_note
                    ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'BOM Based Component Check'

        ws.append(['BOM Component', 'BOM Description', 'MSPEKE Description', 'BOM Qty', 'MSPEKE Notes', 'Hardqualmatrix Check'])

        for key, values in result.items():
            check_value = 'Not in Hardware Qual Matrix' if key in not_in_hqm_list else ''
            ws.append([key] + values + [check_value])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        return send_file(
            output,
            as_attachment=True,
            download_name='BOM_Based_component_error_list.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def highlight_the_color(file_path, error_result_AV):
    wb = load_workbook(file_path)
    ws = wb['Program Matrix']
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color

    for row in ws.iter_rows():
        for cell in row:
            if cell.value in error_result_AV:
                cell.fill = fill

    wb.save('program_matrix_highlight_color.xlsx')

#Smith-Waterman algorithm計算文字相似度
def smith_waterman( seq1, seq2, match_score=2, mismatch_score=-1, gap_score=-1):  
    len1, len2 = len(seq1), len(seq2)
    score_matrix = np.zeros((len1 + 1, len2 + 1), dtype=int)
    traceback_matrix = np.zeros((len1 + 1, len2 + 1), dtype=int)

    max_score = 0
    max_pos = None
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            match = score_matrix[i-1][j-1] + (match_score if seq1[i-1] == seq2[j-1] else mismatch_score)
            delete = score_matrix[i-1][j] + gap_score
            insert = score_matrix[i][j-1] + gap_score
            score_matrix[i][j] = max(0, match, delete, insert)
            if score_matrix[i][j] == 0:
                traceback_matrix[i][j] = 0
            elif score_matrix[i][j] == match:
                traceback_matrix[i][j] = 1
            elif score_matrix[i][j] == delete:
                traceback_matrix[i][j] = 2
            elif score_matrix[i][j] == insert:
                traceback_matrix[i][j] = 3

            if score_matrix[i][j] >= max_score:
                max_score = score_matrix[i][j]
                max_pos = (i, j)
    
    aligned_seq1, aligned_seq2 = [], []
    i, j = max_pos
    while traceback_matrix[i][j] != 0:
        if traceback_matrix[i][j] == 1:
            aligned_seq1.append(seq1[i-1])
            aligned_seq2.append(seq2[j-1])
            i -= 1
            j -= 1
        elif traceback_matrix[i][j] == 2:
            aligned_seq1.append(seq1[i-1])
            aligned_seq2.append('-')
            i -= 1
        elif traceback_matrix[i][j] == 3:
            aligned_seq1.append('-')
            aligned_seq2.append(seq2[j-1])
            j -= 1

    return max_score
if __name__ == '__main__':
    app.run(debug=True, port=5000)
