from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
import os
from io import BytesIO

import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "http://localhost:3000"}})  #讓3000來的請求都通過CORS 之後架server會需要改

UPLOAD_FOLDER = './uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

file_names = {}  #存上傳的檔案 因為上傳的program matrix不會真的只叫做"program matrix" 所以用一個字典來存檔名

@app.route('/', methods=['GET'])
def get_data():
    return jsonify({"message": "Hello from Flask!"})

@app.route('/upload', methods=['POST'])
def upload_file():
    global file_names

    # 獲取上傳的文件
    program_matrix_file = request.files.get('programMatrixFile')
    mspeke_file = request.files.get('mspekeFile')
    hardware_qual_matrix_file = request.files.get('hardwareQualMatrixFile')

    # 檢查是否缺少文件
    if not program_matrix_file or not mspeke_file or not hardware_qual_matrix_file:
        return jsonify({"error": "Missing file"}), 400
    program_matrix_path = os.path.join(UPLOAD_FOLDER, program_matrix_file.filename)
    mspeke_file_path = os.path.join(UPLOAD_FOLDER, mspeke_file.filename)
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, hardware_qual_matrix_file.filename)

    program_matrix_file.save(program_matrix_path)  # 如果檢查通過，保存文件
    mspeke_file.save(mspeke_file_path)
    hardware_qual_matrix_file.save(hardware_qual_matrix_path)

    file_names['program_matrix_file'] = program_matrix_file.filename
    file_names['mspeke_file'] = mspeke_file.filename
    file_names['hardware_qual_matrix_file'] = hardware_qual_matrix_file.filename

    print('1')
    try:
        # 檢查 Program Matrix 文件
        with pd.ExcelFile(program_matrix_file) as xls:
            if 'Program Matrix' not in xls.sheet_names:
                raise ValueError("Sheet 'Program Matrix' not found")
            print('2')    
            df = pd.read_excel(xls, sheet_name='Program Matrix', skiprows=4, usecols="A:I")
            program_matrix_headers = list(df.columns)  # 檢查標題

            expected_program_matrix_headers = [
                'Category / Manufacturing Comments',
                'Release(s)',
                'Description',
                'AV\nLevel 2',
                'Unnamed: 4', 
                'SA\nLevel 3',
                'Unnamed: 6',  
                'Component\nLevel 4',
                'Unnamed: 8'  
            ]

            # 測試 2: 檢查標題是否匹配
            if program_matrix_headers != expected_program_matrix_headers:
                
                raise ValueError("Headers in Program Matrix do not match the expected values.")
            print('3')    
        # 檢查 Mspeke 文件
        with pd.ExcelFile(mspeke_file) as xls:
            if 'HW' not in xls.sheet_names:
                print('4')
                raise ValueError("Sheet 'HW' not found")

            df = pd.read_excel(xls, sheet_name='HW', skiprows=4)
            mspeke_headers = list(df.columns)  # 檢查標題
            if 'Feature Full Name' not in mspeke_headers or 'Notes' not in mspeke_headers:
                
                raise ValueError("'Feature Full Name' or 'Notes' not found in Mspeke headers")
            print('5')    
            print('ssqq4')
        # 檢查 Hardware Qual Matrix 文件
        with pd.ExcelFile(hardware_qual_matrix_file) as xls:
            print('ssqq')
            df = pd.read_excel(xls, skiprows=1)
            hqm_headers = list(df.columns)  # 檢查標題
            print(hqm_headers)
            if 'HP Part No.' not in hqm_headers:
                print('qq')
                raise ValueError("'HP Part No.' not found in HQM headers")
            print('6')
        # 返回成功訊息
        return jsonify({"message": "Files uploaded and validated successfully"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/delete', methods=['POST'])
def delete_folder():
    try:
        if os.path.exists(UPLOAD_FOLDER):
            shutil.rmtree(UPLOAD_FOLDER)  # 刪除整個資料夾及其內容
            os.makedirs(UPLOAD_FOLDER)    # 刪除後重新創建空的資料夾，方便之後的上傳操作
            
            return jsonify({"message": "資料夾已成功刪除"}), 200
        else:
            return jsonify({"error": "資料夾不存在"}), 400
    except Exception as e:
        return jsonify({"error": f"刪除資料夾過程中發生錯誤: {str(e)}"}), 500

@app.route('/bom_cost_check', methods=['GET'])   #目前是回傳只有價格不同的列表 不是有highlight color的
def bom_cost_check():
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))

    if not os.path.exists(program_matrix_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400

    try:
        with pd.ExcelFile(program_matrix_path) as xls:      #使用 with 語句確保文件被關閉!!!! 因為如果不是用with檔案就會一直讓檔案是被操作狀態 會導致之後要按delete時出現 E-busy error
            sheet_names = xls.sheet_names
            df = pd.read_excel(xls, sheet_name=sheet_names[1], skiprows=4)  # sheet_names[1] is Program Matrix
        
        df.rename(columns={'Unnamed: 4': 'av_price', 'Unnamed: 6': 'sa_price'}, inplace=True)
        df['Group'] = df['Release(s)'].ffill()

        block_start_indices = df.index[df['Release(s)'] == 'All (NPI 2024)'].tolist()

        error_result = []

        for start, end in zip(block_start_indices, block_start_indices[1:] + [None]):
            block = df.iloc[start:end]
            
            if 'av_price' in block.columns and 'sa_price' in block.columns:
                total_av_price = block['av_price'].sum()
                total_sa_price = block['sa_price'].sum()
                total_different_price = total_av_price - total_sa_price

                if abs(total_different_price) > 0.01:
                    error_result.append([block.iloc[0]['AV\nLevel 2'], block.iloc[0]['Description'], total_av_price, total_sa_price, total_different_price])
                    
        
        wb = Workbook() #創建工作簿並將結果寫入到指定的工作表中
        ws = wb.active
        ws.title = 'BOM Cost Check'

        ws.append(['AV Level 2', 'Description', 'Total AV Price', 'Total SA Price', 'Price Difference'])    #寫入header

        for row in error_result:    #寫入數據
            ws.append(row)

        
        excel_buffer = BytesIO()    #保存工作簿到BytesIO對象
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()                  #確保工作簿已關閉

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='program_matrix_highlight_color.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/bom_cost_check_for_highlight_file', methods=['GET'])
def bom_cost_check_for_highlight_file():
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))

    if not os.path.exists(program_matrix_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400

    try:
        with pd.ExcelFile(program_matrix_path) as xls:  # 使用 with 确保文件关闭
            sheet_names = xls.sheet_names
            df = pd.read_excel(xls, sheet_name=sheet_names[1], skiprows=4)  # sheet_names[1] is Program Matrix
        
        df.rename(columns={'Unnamed: 4': 'av_price', 'Unnamed: 6': 'sa_price'}, inplace=True)
        df['Group'] = df['Release(s)'].ffill()

        block_start_indices = df.index[df['Release(s)'] == 'All (NPI 2024)'].tolist()

        error_result_AV = []

        for start, end in zip(block_start_indices, block_start_indices[1:] + [None]):
            block = df.iloc[start:end]
            
            if 'av_price' in block.columns and 'sa_price' in block.columns:
                total_av_price = block['av_price'].sum()
                total_sa_price = block['sa_price'].sum()
                total_different_price = total_av_price - total_sa_price

                if abs(total_different_price) > 0.01:
                    error_result_AV.append(block.iloc[0]['AV\nLevel 2'])

        # 加载工作簿并突出显示错误的AV Level 2
        wb = load_workbook(program_matrix_path)
        ws = wb['Program Matrix']
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色高亮

        for row in ws.iter_rows():
            for cell in row:
                if cell.value in error_result_AV:
                    cell.fill = fill

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        wb.close()

        # 将 BytesIO 指针重置到文件的开始
        excel_buffer.seek(0)

        # 返回内存中的 Excel 文件给用户
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='program_matrix_highlight_color.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/hqm_based_component_check', methods=['GET'])
def hqm_based_component_check():
    global file_names
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))
    mspeke_path = os.path.join(UPLOAD_FOLDER, file_names.get('mspeke_file', ''))
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('hardware_qual_matrix_file', ''))
    print(file_names)

    if not (os.path.exists(program_matrix_path)):
        print(f"Error: {program_matrix_path} 不存在")
    if not (os.path.exists(mspeke_path)):
        print(f"Error: {mspeke_path} 不存在")
    if not (os.path.exists(hardware_qual_matrix_path)):
        print(f"Error: {hardware_qual_matrix_path} 不存在")
    
    try:
        with pd.ExcelFile(program_matrix_path) as pm, pd.ExcelFile(mspeke_path) as mspeke:
            pm_sheet_names = pm.sheet_names
            mspeke_sheet_names = mspeke.sheet_names

            df_pm = pd.read_excel(pm, sheet_name=pm_sheet_names[1], skiprows=4)
            df_mspeke = pd.read_excel(mspeke, sheet_name=mspeke_sheet_names[1], skiprows=4)
            df_hqm = pd.read_excel(hardware_qual_matrix_path, skiprows=1, usecols=['HP Part No.', 'Qual Status'])
        
        df_mspeke = df_mspeke.iloc[:, [1, 4, 8]].dropna(subset=[df_mspeke.columns[4]])
        df_hqm = df_hqm.dropna(subset=['HP Part No.', 'Qual Status'])

        sections = {}
        current_label = None
        result = {}

        for index, row in df_pm.iterrows():
            if pd.notna(row['SA\nLevel 3']):
                current_label = row['SA\nLevel 3']
                sections[current_label] = []
            if current_label:
                sections[current_label].append(row)

        for label in sections:
            sections[label] = pd.DataFrame(sections[label])

        for index, hqm_row in df_hqm.iterrows():
            hqm_part_number = hqm_row['HP Part No.']
            hqm_qual_status = hqm_row['Qual Status']
            result[hqm_part_number] = [hqm_qual_status]
            HAS_FOUNDED = False
            for key, value in sections.items():
                if hqm_part_number in value['Component\nLevel 4'].values:
                    HAS_FOUNDED = True
                    component_value = value[value['Component\nLevel 4'] == hqm_part_number]
                    component_description = component_value['Description'].values[0]
                    component_qty = value.iloc[0]['Quantity']

                    result[hqm_part_number].extend([component_qty, component_description])

            if not HAS_FOUNDED:
                result[hqm_part_number].append('Cannot find this component in Program Matrix')

        #result存了 key:hqm part num, value: ['Qual Status', 'BOM qty, 'BOM desceiption']
        final_result = []
        smith_waterman_THRESHOLD = 0
        for key, value in result.items():
            max_ratio = 0
            max_mspeke_item = None
            if len(value) > 2: #代表同時有出現在hqm bom的
                for idx, mspeke_item in df_mspeke.iterrows():
                    ratio = smith_waterman(value[2], mspeke_item['Feature Full Name'])
                    if max_ratio < ratio:
                        max_ratio = ratio
                        max_mspeke_item = mspeke_item
                final_result.append([
                    key, value[0], value[2], value[1], max_mspeke_item['Feature Full Name'], max_mspeke_item['Notes']
                ])
            else:
                final_result.append([key, value[0], value[1]])

        wb = Workbook()
        ws = wb.active
        ws.title = 'HQM Based Component Check'

        ws.append([
            'Component Part Number',
            'Qual Status', 'Program Matrix Description', 'Program Matrix Qty', 'BOM Description -> MSPEKE Description', 'BOM Description -> MSPEKE Notes'
        ])

        for row in final_result:
            ws.append(row)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='HQM_Based_component_error_list.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route('/bom_based_component_check', methods=['GET'])
def bom_based_component_check():
    global file_names
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))
    mspeke_path = os.path.join(UPLOAD_FOLDER, file_names.get('mspeke_file', ''))
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('hardware_qual_matrix_file', ''))

    if not (os.path.exists(program_matrix_path) and os.path.exists(mspeke_path) and os.path.exists(hardware_qual_matrix_path)):
        return jsonify({"error": "One or more files are missing. Please upload the files first."}), 400

    try:
        with pd.ExcelFile(program_matrix_path) as pm, pd.ExcelFile(mspeke_path) as mspeke:
            pm_sheet_names = pm.sheet_names
            mspeke_sheet_names = mspeke.sheet_names

            df_pm = pd.read_excel(pm, sheet_name=pm_sheet_names[1], skiprows=4)
            df_mspeke = pd.read_excel(mspeke, sheet_name=mspeke_sheet_names[1], skiprows=4)
            df_hqm = pd.read_excel(hardware_qual_matrix_path, skiprows=1)

        start_idx = df_pm[df_pm['Category / Manufacturing Comments'] == 'DIB Hardware'].index[0]
        end_idx = df_pm[(df_pm.index > start_idx) & (df_pm['Category / Manufacturing Comments'].notna())].index[0]
        df_pm = df_pm.drop(df_pm.index[start_idx:end_idx]).reset_index(drop=True)

        df_mspeke = df_mspeke[['Feature\nID', 'Feature Full Name', 'Notes']].dropna(subset=['Notes'])
        df_hqm = df_hqm['HP Part No.'].dropna()

        sections = {}
        not_in_hqm_list = []
        result = {}

        current_label = None
        for index, row in df_pm.iterrows():
            if pd.notna(row['SA\nLevel 3']):
                current_label = row['SA\nLevel 3']
                sections[current_label] = []
            if current_label:
                sections[current_label].append(row)

        for label in sections:
            sections[label] = pd.DataFrame(sections[label])

        for key, value in sections.items():
            for _, row in value.iterrows():
                if pd.notna(row['Component\nLevel 4']):
                    if row['Component\nLevel 4'] not in df_hqm.values:
                        not_in_hqm_list.append(row['Component\nLevel 4'])

                    max_ratio = 0
                    max_mspeke_item = None
                    max_mspeke_item_note = None
                    for _, item in df_mspeke.iterrows():
                        if pd.notna(item['Feature Full Name']):
                            ratio = smith_waterman(row['Description'], item['Feature Full Name'])
                            if ratio > max_ratio:
                                max_ratio = ratio
                                max_mspeke_item = item['Feature Full Name']
                                max_mspeke_item_note = item['Notes']

                    result[row['Component\nLevel 4']] = [
                        row['Description'],
                        max_mspeke_item,
                        value.iloc[0]['Quantity'],
                        max_mspeke_item_note
                    ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'BOM Based Component Check'

        ws.append(['BOM Component', 'BOM Description', 'MSPEKE Description', 'BOM Qty', 'MSPEKE Notes', 'Hardqualmatrix Check'])

        for key, values in result.items():
            check_value = 'Not in Hardware Qual Matrix' if key in not_in_hqm_list else ''
            ws.append([key] + values + [check_value])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        return send_file(
            output,
            as_attachment=True,
            download_name='BOM_Based_component_error_list.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

#Smith-Waterman algorithm計算文字相似度
def smith_waterman( seq1, seq2, match_score=2, mismatch_score=-1, gap_score=-1):  
    len1, len2 = len(seq1), len(seq2)
    score_matrix = np.zeros((len1 + 1, len2 + 1), dtype=int)
    traceback_matrix = np.zeros((len1 + 1, len2 + 1), dtype=int)

    max_score = 0
    max_pos = None
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            match = score_matrix[i-1][j-1] + (match_score if seq1[i-1] == seq2[j-1] else mismatch_score)
            delete = score_matrix[i-1][j] + gap_score
            insert = score_matrix[i][j-1] + gap_score
            score_matrix[i][j] = max(0, match, delete, insert)
            if score_matrix[i][j] == 0:
                traceback_matrix[i][j] = 0
            elif score_matrix[i][j] == match:
                traceback_matrix[i][j] = 1
            elif score_matrix[i][j] == delete:
                traceback_matrix[i][j] = 2
            elif score_matrix[i][j] == insert:
                traceback_matrix[i][j] = 3

            if score_matrix[i][j] >= max_score:
                max_score = score_matrix[i][j]
                max_pos = (i, j)
    
    aligned_seq1, aligned_seq2 = [], []
    i, j = max_pos
    while traceback_matrix[i][j] != 0:
        if traceback_matrix[i][j] == 1:
            aligned_seq1.append(seq1[i-1])
            aligned_seq2.append(seq2[j-1])
            i -= 1
            j -= 1
        elif traceback_matrix[i][j] == 2:
            aligned_seq1.append(seq1[i-1])
            aligned_seq2.append('-')
            i -= 1
        elif traceback_matrix[i][j] == 3:
            aligned_seq1.append('-')
            aligned_seq2.append(seq2[j-1])
            j -= 1

    return max_score

if __name__ == '__main__':
    app.run(debug=True, port=5000)
