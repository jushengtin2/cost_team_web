from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import pandas as pd
import numpy as np
import os
from io import BytesIO
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "http://localhost:3000"}})  #讓3000來的請求都通過CORS 之後架server會需要改

UPLOAD_FOLDER = './uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

file_names = {}  #存上傳的檔案 因為上傳的program matrix不會真的只叫做"program matrix" 所以用一個字典來存檔名

@app.route('/', methods=['GET'])
def get_data():
    return jsonify({"message": "Hello from Flask!"})

@app.route('/DT_upload', methods=['POST'])
def upload_file():
    global file_names

    # 獲取上傳的文件
    program_matrix_file = request.files.get('programMatrixFile')
    mspeke_file = request.files.get('mspekeFile')
    hardware_qual_matrix_file = request.files.get('hardwareQualMatrixFile')

    # 檢查是否缺少文件
    if not program_matrix_file or not mspeke_file or not hardware_qual_matrix_file:
        return jsonify({"error": "Missing file"}), 400
    program_matrix_path = os.path.join(UPLOAD_FOLDER, program_matrix_file.filename)
    mspeke_file_path = os.path.join(UPLOAD_FOLDER, mspeke_file.filename)
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, hardware_qual_matrix_file.filename)

    program_matrix_file.save(program_matrix_path)  # 如果檢查通過，保存文件
    mspeke_file.save(mspeke_file_path)
    hardware_qual_matrix_file.save(hardware_qual_matrix_path)

    file_names['program_matrix_file'] = program_matrix_file.filename
    file_names['mspeke_file'] = mspeke_file.filename
    file_names['hardware_qual_matrix_file'] = hardware_qual_matrix_file.filename

    try:
        #檢查 Program Matrix 文件
        with pd.ExcelFile(program_matrix_file) as xls:
            expected_program_matrix_headers = [
                'Category / Manufacturing Comments',
                'Release(s)',
                'Description',
                'AV\nLevel 2',
                'Unnamed: 4', 
                'SA\nLevel 3',
                'Unnamed: 6',  
                'Component\nLevel 4',
                'Unnamed: 8'  
            ]
            #DT有四個sheet每個都要看是不是格式正確
            for sheet_name in xls.sheet_names:
                print(f'Processing sheet: {sheet_name}')    
                df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=4, usecols="A:I") #我設定是檢查前A到Icolumn標題就好
                program_matrix_headers = list(df.columns)  

                # 測試 2: 檢查標題是否匹配
                if program_matrix_headers != expected_program_matrix_headers:
                    raise ValueError(f"Headers in sheet '{sheet_name}' do not match the expected values.")
                print('Headers match expected values.') 

        # 檢查 Mspeke 文件
        with pd.ExcelFile(mspeke_file) as xls:
            if 'HW' not in xls.sheet_names:
                print('4')
                raise ValueError("Sheet 'HW' not found")

            df = pd.read_excel(xls, sheet_name='HW', skiprows=4)
            mspeke_headers = list(df.columns)  # 檢查標題
            if 'Feature Full Name' not in mspeke_headers or 'Notes' not in mspeke_headers:
                
                raise ValueError("'Feature Full Name' or 'Notes' not found in Mspeke headers")
            print('5')    
            print('ssqq4')
        # 檢查 Hardware Qual Matrix 文件
        with pd.ExcelFile(hardware_qual_matrix_file) as xls:
            print('ssqq')
            df = pd.read_excel(xls, skiprows=1)
            hqm_headers = list(df.columns)  # 檢查標題
            print(hqm_headers)
            if 'HP Part No.' not in hqm_headers:
                print('qq')
                raise ValueError("'HP Part No.' not found in HQM headers")
            print('6')
        # 返回成功訊息
        return jsonify({"message": "Files uploaded and validated successfully"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/DT_delete', methods=['POST'])
def delete_folder():
    try:
        if os.path.exists(UPLOAD_FOLDER):
            shutil.rmtree(UPLOAD_FOLDER)  # 刪除整個資料夾及其內容
            os.makedirs(UPLOAD_FOLDER)    # 刪除後重新創建空的資料夾，方便之後的上傳操作
            
            return jsonify({"message": "資料夾已成功刪除"}), 200
        else:
            return jsonify({"error": "資料夾不存在"}), 400
    except Exception as e:
        return jsonify({"error": f"刪除資料夾過程中發生錯誤: {str(e)}"}), 500

@app.route('/DT_bom_cost_check', methods=['GET'])
def bom_cost_check():
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))

    if not os.path.exists(program_matrix_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400

    try:
        with pd.ExcelFile(program_matrix_path) as xls:
            sheet_names = xls.sheet_names
            
            # 創建一個新的工作簿來存儲每個工作表的結果
            wb = Workbook()
            ws = wb.active
            ws.title = 'BOM Cost Check'  # 為第一個工作表命名
            
            for i, sheet_name in enumerate(sheet_names):
                df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=4)
                df.rename(columns={'Unnamed: 4': 'av_price', 'Unnamed: 6': 'sa_price'}, inplace=True)
                df['Group'] = df['Release(s)'].ffill()

                block_start_indices = df.index[df['Release(s)'].notna()].tolist()
                error_result = []

                for start, end in zip(block_start_indices, block_start_indices[1:] + [None]):
                    block = df.iloc[start:end]
                    
                    if 'av_price' in block.columns and 'sa_price' in block.columns:
                        total_av_price = block['av_price'].sum()
                        total_sa_price = block['sa_price'].sum()
                        total_different_price = total_av_price - total_sa_price

                        if abs(total_different_price) > 0.01:
                            error_result.append([block.iloc[0]['AV\nLevel 2'], block.iloc[0]['Description'], total_av_price, total_sa_price, total_different_price])
                
                # 新增一個工作表來存儲當前 sheet 的結果
                if i == 0:
                    ws = wb.active  # First sheet is already created, just use it
                else:
                    ws = wb.create_sheet(title=f'Sheet {i+1}')
                
                # 寫入結果到工作表
                ws.append(['AV Level 2', 'Description', 'Total AV Price', 'Total SA Price', 'Price Difference'])
                for row in error_result:
                    ws.append(row)

        # 將工作簿保存到 BytesIO 對象
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='program_matrix_highlight_color.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/DT_bom_cost_check_for_highlight_file', methods=['GET'])
def bom_cost_check_for_highlight_file():
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))

    if not os.path.exists(program_matrix_path):
        return jsonify({"error": "The program matrix file is missing. Please upload the file first."}), 400

    try:
        error_results = {}  #存每個sheet的結果

        with pd.ExcelFile(program_matrix_path) as xls:
            sheet_names = xls.sheet_names

            for sheet_name in sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=4)
                df.rename(columns={'Unnamed: 4': 'av_price', 'Unnamed: 6': 'sa_price'}, inplace=True)
                df['Group'] = df['Release(s)'].ffill()

                block_start_indices = df.index[df['Release(s)'].notna()].tolist()

                error_result_AV = []

                for start, end in zip(block_start_indices, block_start_indices[1:] + [None]):
                    block = df.iloc[start:end]
                    
                    if 'av_price' in block.columns and 'sa_price' in block.columns:
                        total_av_price = block['av_price'].sum()
                        total_sa_price = block['sa_price'].sum()
                        total_different_price = total_av_price - total_sa_price

                        if abs(total_different_price) > 0.01:
                            error_result_AV.append(block.iloc[0]['AV\nLevel 2'])
                
                error_results[sheet_name] = error_result_AV

        wb = Workbook()
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色

        for sheet_name, errors in error_results.items():
            ws = wb.create_sheet(title=f'{sheet_name} Errors')
            ws.append(['AV Level 2', 'Description', 'Total AV Price', 'Total SA Price', 'Price Difference'])

            df = pd.read_excel(program_matrix_path, sheet_name=sheet_name, skiprows=4)
            df.rename(columns={'Unnamed: 4': 'av_price', 'Unnamed: 6': 'sa_price'}, inplace=True)
            df['Group'] = df['Release(s)'].ffill()

            block_start_indices = df.index[df['Release(s)'].notna()].tolist()

            for start, end in zip(block_start_indices, block_start_indices[1:] + [None]):
                block = df.iloc[start:end]
                
                if 'av_price' in block.columns and 'sa_price' in block.columns:
                    total_av_price = block['av_price'].sum()
                    total_sa_price = block['sa_price'].sum()
                    total_different_price = total_av_price - total_sa_price

                    if abs(total_different_price) > 0.01:
                        av_level_2 = block.iloc[0]['AV\nLevel 2']
                        row_data = [av_level_2, block.iloc[0]['Description'], total_av_price, total_sa_price, total_different_price]
                        ws.append(row_data)

                        # highlight error的 AV Level 2
                        for row in ws.iter_rows(min_row=2, max_col=1):  #從第二row開始因為第一row是標題
                            for cell in row:
                                if cell.value == av_level_2:
                                    cell.fill = fill

        # 移除默认生成的空白工作表
        if 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        wb.close()

        # 将 BytesIO 指针重置到文件的开始
        excel_buffer.seek(0)

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name='program_matrix_highlight_color.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/DT_hqm_based_component_check', methods=['GET'])
def hqm_based_component_check():
    global file_names
    
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))
    mspeke_path = os.path.join(UPLOAD_FOLDER, file_names.get('mspeke_file', ''))
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('hardware_qual_matrix_file', ''))
    print(file_names)

    if not (os.path.exists(program_matrix_path)):
        print(f"Error: {program_matrix_path} 不存在")
    if not (os.path.exists(mspeke_path)):
        print(f"Error: {mspeke_path} 不存在")
    if not (os.path.exists(hardware_qual_matrix_path)):
        print(f"Error: {hardware_qual_matrix_path} 不存在")
    
    try:
        with pd.ExcelFile(program_matrix_path) as pm, pd.ExcelFile(mspeke_path) as mspeke, pd.ExcelFile(hardware_qual_matrix_path) as hqm:
            pm_sheet_names = pm.sheet_names
            mspeke_sheet_names = mspeke.sheet_names

            df_pm = pd.read_excel(pm, sheet_name=pm_sheet_names[1], skiprows=4)
            df_mspeke = pd.read_excel(mspeke, sheet_name=mspeke_sheet_names[1], skiprows=4)
            df_hqm = pd.read_excel(hqm, skiprows=1, usecols=['ID','HP Part No.', 'Qual Status'])

            start_idx = df_pm[df_pm['Category / Manufacturing Comments'] == 'DIB Hardware'].index[0]
            end_idx = df_pm[(df_pm.index > start_idx) & (df_pm['Category / Manufacturing Comments'].notna())].index[0]

            df_pm = df_pm.drop(df_pm.index[start_idx:end_idx]).reset_index(drop=True)
            df_mspeke = df_mspeke.iloc[:, [1, 4, 8]].dropna(subset=[df_mspeke.columns[4]])
            
            sections = {}
            current_label = None

            for index, row in df_pm.iterrows():
                if pd.notna(row['SA\nLevel 3']):
                    current_label = row['SA\nLevel 3']
                    sections[current_label] = []
                if current_label:
                    sections[current_label].append(row)

            for label in sections:
                sections[label] = pd.DataFrame(sections[label])

            df_hqm['BOM Description'] = None
            df_hqm['MSPEKE Description'] = None

            for index, row in df_hqm.iterrows():
                if row['HP Part No.'] and row['Qual Status']:
                    hqm_part_number = row['HP Part No.']
                    hqm_qual_status = row['Qual Status']
                    for key, value in sections.items():
                        if hqm_part_number in value['Component\nLevel 4'].values:
                            component_value = value[value['Component\nLevel 4'] == hqm_part_number]
                            component_description = component_value['Description'].values[0]
                            df_hqm.at[index, 'BOM Description'] = component_description
                            
                            max_ratio = 0
                            max_mspeke_item = None
                            for idx, mspeke_item in df_mspeke.iterrows():
                                ratio = smith_waterman(component_description, mspeke_item['Feature Full Name'])
                                if max_ratio < ratio:
                                    max_ratio = ratio
                                    max_mspeke_item = mspeke_item['Feature Full Name']
                            df_hqm.at[index, 'MSPEKE Description'] = max_mspeke_item
                            break

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_hqm.to_excel(writer, index=False, sheet_name='Component Error List')

            # 獲取當前活頁簿和工作表
            workbook = writer.book
            worksheet = workbook.active

            # 定義綠色和淺咖啡色填充樣式
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            light_brown_fill = PatternFill(start_color="D3B58C", end_color="D3B58C", fill_type="solid")

            # 遍歷 df_hqm 的每一行，根據條件設置顏色
            for row_idx, row in enumerate(df_hqm.itertuples(), start=2):  # 開始於第2行（Excel中的行從1開始）
                if pd.isna(row._2):  # '_2' 對應的是 'HP Part No.' 列
                    if pd.notna(row.ID) and not any(char.isdigit() for char in str(row.ID)):
                        # ID 欄位沒有數字時，設置綠色
                        for col_idx in range(1, len(df_hqm.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = green_fill
                    else:
                        # ID 欄位包含數字時，設置淺咖啡色
                        for col_idx in range(1, len(df_hqm.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = light_brown_fill

        output.seek(0)

        # 返回文件作為下載
        return send_file(
            output,
            as_attachment=True,
            download_name='HQM_Based_component_error_list.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route('/DT_bom_based_component_check', methods=['GET'])
def bom_based_component_check():
    global file_names
    program_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('program_matrix_file', ''))
    mspeke_path = os.path.join(UPLOAD_FOLDER, file_names.get('mspeke_file', ''))
    hardware_qual_matrix_path = os.path.join(UPLOAD_FOLDER, file_names.get('hardware_qual_matrix_file', ''))

    if not (os.path.exists(program_matrix_path) and os.path.exists(mspeke_path) and os.path.exists(hardware_qual_matrix_path)):
        return jsonify({"error": "One or more files are missing. Please upload the files first."}), 400
    
    try:
        # 讀取 Excel 文件
        with pd.ExcelFile(program_matrix_path) as pm, pd.ExcelFile(mspeke_path) as mspeke, pd.ExcelFile(hardware_qual_matrix_path) as hqm:
            pm_sheet_names = pm.sheet_names
            mspeke_sheet_names = mspeke.sheet_names

            df_pm = pd.read_excel(pm, sheet_name=pm_sheet_names[1], skiprows=4)
            df_mspeke = pd.read_excel(mspeke, sheet_name=mspeke_sheet_names[1], skiprows=4)
            df_hqm = pd.read_excel(hqm, skiprows=1)

            # 清理和準備數據
            start_idx = df_pm[df_pm['Category / Manufacturing Comments'] == 'DIB Hardware'].index[0]  # 去除DIB部分
            end_idx = df_pm[(df_pm.index > start_idx) & (df_pm['Category / Manufacturing Comments'].notna())].index[0]

            df_pm = df_pm.drop(df_pm.index[start_idx:end_idx]).reset_index(drop=True)
            df_mspeke = df_mspeke.iloc[:, [1, 4, 8]].dropna(subset=[df_mspeke.columns[1]])
            df_hqm = df_hqm[['HP Part No.', 'Qual Status']].dropna(subset=['HP Part No.'])
            hqm_dict = df_hqm.set_index('HP Part No.')['Qual Status'].to_dict()

            # 創建新列來存儲結果
            df_pm['Max_MSPEKE_Item'] = None
            df_pm['Max_MSPEKE_Item_Note'] = None
            df_pm['HQM status'] = None

            for index, row in df_pm.iterrows():
                if pd.notna(row['Component\nLevel 4']):
                    cmp_level4 = row['Description']
                    if 'lbl' not in cmp_level4.lower() and 'doc' not in cmp_level4.lower() and 'icon' not in cmp_level4.lower(): #把不需要對照的刪掉
                        max_ratio = 0
                        max_mspeke_item = None
                        max_mspeke_item_note = None

                        for _, item in df_mspeke.iterrows():
                            ratio = smith_waterman(cmp_level4, item['Feature Full Name'])
                            if ratio > max_ratio:
                                max_ratio = ratio
                                max_mspeke_item = item['Feature Full Name']
                                max_mspeke_item_note = item['Notes']
                            

                        # 插入結果到對應的行
                        df_pm.at[index, 'Max_MSPEKE_Item'] = max_mspeke_item
                        df_pm.at[index, 'Max_MSPEKE_Item_Note'] = max_mspeke_item_note
                        print(f'1. {cmp_level4}, 2.{max_mspeke_item} , 3. {max_ratio}')

                        if row['Component\nLevel 4'] not in hqm_dict:
                            df_pm.at[index, 'HQM status'] = 'Not in Hardware Qual Matrix'
                        else:
                            qual_status_value = hqm_dict[row['Component\nLevel 4']]
                            df_pm.at[index, 'HQM status'] = qual_status_value

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_pm.to_excel(writer, index=False, sheet_name='BOM Based Component Check')

            # 獲取工作表
            worksheet = writer.sheets['BOM Based Component Check']

            # 定義顏色
            pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 藍色填充
            header_yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 標題黃色填充

            # 定義邊框樣式
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # 使用列名來填充顏色
            max_mspeke_item_col = df_pm.columns.get_loc('Max_MSPEKE_Item') + 1  # +1 是因為 openpyxl 列索引從 1 開始
            max_mspeke_item_note_col = df_pm.columns.get_loc('Max_MSPEKE_Item_Note') + 1
            hqm_status_col = df_pm.columns.get_loc('HQM status') + 1

            # 先將不包含特定列的標題行（第一行）塗上黃色
            for col_idx in range(1, worksheet.max_column + 1):
                if col_idx not in [max_mspeke_item_col, max_mspeke_item_note_col, hqm_status_col]:
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_yellow_fill
                    cell.border = thin_border

            # 塗上藍色
            for row_idx in range(2, worksheet.max_row + 1):  # 從第2行開始（跳過標題行）
                category_comments = worksheet.cell(row=row_idx, column=df_pm.columns.get_loc('Category / Manufacturing Comments') + 1).value
                release_s = worksheet.cell(row=row_idx, column=df_pm.columns.get_loc('Release(s)') + 1).value
                description = worksheet.cell(row=row_idx, column=df_pm.columns.get_loc('Description') + 1).value

                if category_comments and (release_s is None or release_s == '') and (description is None or description == ''):
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = blue_fill
                        cell.border = thin_border
            # 塗上粉紅色和黃色並保持邊框
            for cell in worksheet[worksheet.cell(row=1, column=max_mspeke_item_col).column_letter]:
                cell.fill = pink_fill
                cell.border = thin_border

            for cell in worksheet[worksheet.cell(row=1, column=max_mspeke_item_note_col).column_letter]:
                cell.fill = pink_fill
                cell.border = thin_border

            for cell in worksheet[worksheet.cell(row=1, column=hqm_status_col).column_letter]:
                cell.fill = yellow_fill
                cell.border = thin_border

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='BOM_Based_component_error_list.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    

#Smith-Waterman algorithm計算文字相似度
def smith_waterman( seq1, seq2, match_score=2, mismatch_score=-1, gap_score=-1):  
    len1, len2 = len(seq1), len(seq2)
    score_matrix = np.zeros((len1 + 1, len2 + 1), dtype=int)
    traceback_matrix = np.zeros((len1 + 1, len2 + 1), dtype=int)

    max_score = 0
    max_pos = None
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            match = score_matrix[i-1][j-1] + (match_score if seq1[i-1] == seq2[j-1] else mismatch_score)
            delete = score_matrix[i-1][j] + gap_score
            insert = score_matrix[i][j-1] + gap_score
            score_matrix[i][j] = max(0, match, delete, insert)
            if score_matrix[i][j] == 0:
                traceback_matrix[i][j] = 0
            elif score_matrix[i][j] == match:
                traceback_matrix[i][j] = 1
            elif score_matrix[i][j] == delete:
                traceback_matrix[i][j] = 2
            elif score_matrix[i][j] == insert:
                traceback_matrix[i][j] = 3

            if score_matrix[i][j] >= max_score:
                max_score = score_matrix[i][j]
                max_pos = (i, j)
    
    aligned_seq1, aligned_seq2 = [], []
    i, j = max_pos
    while traceback_matrix[i][j] != 0:
        if traceback_matrix[i][j] == 1:
            aligned_seq1.append(seq1[i-1])
            aligned_seq2.append(seq2[j-1])
            i -= 1
            j -= 1
        elif traceback_matrix[i][j] == 2:
            aligned_seq1.append(seq1[i-1])
            aligned_seq2.append('-')
            i -= 1
        elif traceback_matrix[i][j] == 3:
            aligned_seq1.append('-')
            aligned_seq2.append(seq2[j-1])
            j -= 1

    return max_score




if __name__ == '__main__':
    app.run(debug=True, port=5000)
